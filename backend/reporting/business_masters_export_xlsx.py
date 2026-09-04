# ====================================
# BUSINESS MASTERS "EXPORT CURRENT TAB" - STYLED .XLSX
# Reuses the exact same visual design as the Customer 360 export
# (backend/reporting/xlsx_style.py) - title banner, blue subtitle
# band, styled header row, bordered grid - in place of the old plain,
# unstyled client-side SheetJS build (the free/community `xlsx`
# package can't apply cell fills/fonts at all, the same limitation
# that originally forced Customer 360 server-side).
#
# Pure presentation change - the data itself is exactly what
# business_masters_export_api.py already computed and returned as raw
# JSON before this: one real sheet per backing table, every DB
# column, in the same order. Only how it's rendered has changed.
# ====================================

import json

from datetime import date, datetime
from decimal import Decimal
from io import BytesIO

from openpyxl import Workbook

from backend.reporting.xlsx_style import (
    title_row,
    subtitle_row,
    table_header_row,
    table_data_row,
    sheet_name as safe_sheet_name
)


# ====================================
# CELL VALUE NORMALIZATION
# Mirrors exactly what a JSON round-trip through the old endpoint
# already did to these same raw ORM values before the frontend's own
# sanitizeRowsForExport() flattened them further - dates/decimals
# stringify the same way FastAPI's jsonable_encoder already stringified
# them, and JSONB list/dict columns flatten the same way that helper
# already flattened them (join / JSON.stringify) - so the actual
# content shown in a cell is unchanged, only its formatting/borders are.
# ====================================

def _cell_value(value):

    if value is None:
        return None

    if isinstance(value, (datetime, date)):
        return value.isoformat()

    if isinstance(value, Decimal):
        return float(value)

    if isinstance(value, (list, tuple)):
        return ", ".join(str(item) for item in value)

    if isinstance(value, dict):
        return json.dumps(value)

    return value


# ====================================
# COLUMN WIDTH
# Sized off the real header/content lengths for that column, not a
# single fixed width - a "name" column and a "created_at" column don't
# read well at the same width. Capped both ends so one very long value
# can't blow the sheet out, and a one-character header doesn't collapse.
# ====================================

def _column_width(header, sample_values):

    longest = len(str(header))

    for value in sample_values:

        if value is not None:
            longest = max(longest, len(str(value)))

    return max(12, min(longest + 2, 45))


# ====================================
# ONE SHEET PER BACKING TABLE
# ====================================

def _build_sheet(wb, tab_label, sheet):

    ws = wb.create_sheet(safe_sheet_name(sheet["name"]))

    rows = sheet["rows"]

    headers = list(rows[0].keys()) if rows else []

    total_cols = max(len(headers), 1)

    row = 1

    title_row(ws, row, total_cols, "JANYU TECHNOLOGIES")
    row += 1

    subtitle_row(ws, row, total_cols, f"Business Masters — {sheet['name']}")
    row += 1

    if not rows:

        ws.cell(row=row, column=1, value="No rows yet.")
        ws.column_dimensions["A"].width = 30

        return ws

    table_header_row(ws, row, headers)
    header_row = row
    row += 1

    normalized_rows = [
        [_cell_value(record.get(header)) for header in headers]
        for record in rows
    ]

    for values in normalized_rows:
        table_data_row(ws, row, values)
        row += 1

    for i, header in enumerate(headers):

        sample = [values[i] for values in normalized_rows[:200]]

        ws.column_dimensions[ws.cell(row=header_row, column=i + 1).column_letter].width = (
            _column_width(header, sample)
        )

    ws.freeze_panes = ws.cell(row=header_row + 1, column=1).coordinate

    return ws


# ====================================
# BUILD WORKBOOK
# ====================================

def build_tab_export_workbook_bytes(tab_label, sheets):

    wb = Workbook()
    wb.remove(wb.active)

    for sheet in sheets:
        _build_sheet(wb, tab_label, sheet)

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    return buffer


# ====================================
# CUSTOMERS TAB'S OWN "EXPORT CURRENT TAB"
# The Customers tab's export is a computed report (customer_master_
# service.py::build_customers_report), not a raw table dump - each of
# its 3 sheets uses display labels rather than raw column names, so
# they're spelled out explicitly here rather than derived from a
# dict's own keys, matching exactly what the old client-side build
# already showed.
# ====================================

_CUSTOMERS_SUMMARY_COLUMNS = [
    ("company", "Company"),
    ("industry", "Industry"),
    ("location", "Location"),
    ("account_manager", "Account Manager"),
    ("total_enquiries", "Total Enquiries"),
    ("total_closed_jobs", "Total Closed Jobs"),
    ("invoice_value", "Invoice Value")
]

_CUSTOMERS_ASSETS_COLUMNS = [
    ("company_name", "Company Name"),
    ("asset_name", "Asset Name"),
    ("closed_jobs_count", "Closed Jobs till date (Count)"),
    ("open_enquiries_count", "Open Enquiries (Till PO Received)"),
    ("enquiry_stage", "Enquiry Stage"),
    ("last_closed_job_date", "Last Closed Job Date"),
    ("next_follow_up_date", "Next Follow-up Date"),
    ("invoice_value", "Invoice Value"),
    ("account_manager", "Account Manager")
]

_CUSTOMERS_CONTACTS_COLUMNS = [
    ("company_name", "Company Name"),
    ("category", "Category"),
    ("industry", "Industry"),
    ("region", "Region"),
    ("gst_number", "GST Number"),
    ("account_manager", "Account Manager"),
    ("poc_name", "POC Name"),
    ("poc_designation", "POC Designation"),
    ("poc_email", "POC Email"),
    ("poc_phone", "POC Phone")
]


def _build_labeled_sheet(wb, sheet_title, columns, rows):

    ws = wb.create_sheet(safe_sheet_name(sheet_title))

    headers = [label for _key, label in columns]
    total_cols = max(len(headers), 1)

    row = 1

    title_row(ws, row, total_cols, "JANYU TECHNOLOGIES")
    row += 1

    subtitle_row(ws, row, total_cols, f"Business Masters — {sheet_title}")
    row += 1

    if not rows:

        ws.cell(row=row, column=1, value="No rows yet.")
        ws.column_dimensions["A"].width = 30

        return ws

    table_header_row(ws, row, headers)
    header_row = row
    row += 1

    normalized_rows = [
        [_cell_value(record.get(key)) for key, _label in columns]
        for record in rows
    ]

    for values in normalized_rows:
        table_data_row(ws, row, values)
        row += 1

    for i, header in enumerate(headers):

        sample = [values[i] for values in normalized_rows[:200]]

        ws.column_dimensions[ws.cell(row=header_row, column=i + 1).column_letter].width = (
            _column_width(header, sample)
        )

    ws.freeze_panes = ws.cell(row=header_row + 1, column=1).coordinate

    return ws


def build_customers_report_workbook_bytes(report):

    wb = Workbook()
    wb.remove(wb.active)

    _build_labeled_sheet(wb, "Customer Summary", _CUSTOMERS_SUMMARY_COLUMNS, report["summary"])
    _build_labeled_sheet(wb, "Assets", _CUSTOMERS_ASSETS_COLUMNS, report["assets"])
    _build_labeled_sheet(wb, "Company & POC", _CUSTOMERS_CONTACTS_COLUMNS, report["contacts"])

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    return buffer
