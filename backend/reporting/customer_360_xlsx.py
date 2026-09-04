# ====================================
# IMPORTS
# ====================================

from io import BytesIO

from openpyxl import Workbook

from backend.services.customer_master_service import get_customer_detail_request

from backend.reporting.xlsx_style import (
    LABEL_COL_WIDTH,
    VALUE_COL_WIDTH,
    sheet_name as _sheet_name,
    title_row as _title_row,
    subtitle_row as _subtitle_row,
    section_band as _section_band,
    field_row as _field_row
)


# ====================================
# FIELD DEFINITIONS
# Mirrors frontend/src/data/surveyProfileFields.js exactly (same keys,
# labels, groups) - kept as a separate Python list rather than shared
# code since the two run in different languages, but must stay in sync
# by hand if either changes.
# ====================================

def _fmt_dt(value):

    if not value:
        return ""

    try:
        return value.strftime("%d-%b-%Y %I:%M %p")
    except AttributeError:
        return str(value)


def _fmt_bool(value):

    if value is None:
        return ""

    return "Yes" if value else "No"


CUSTOMER_FIELDS = [
    ("ID", lambda c: c["id"]),
    ("Company", lambda c: c["company_name"]),
    ("Category", lambda c: c["category"] or ""),
    ("Industry", lambda c: c["industry"] or ""),
    ("Region", lambda c: c["region"] or ""),
    ("Account Owner", lambda c: c["owner_name"] or c["owner"] or ""),
    ("Created By", lambda c: c["created_by_name"] or ""),
    ("GST Number", lambda c: c["gst_number"] or ""),
    ("Next Follow-up Date", lambda c: c["next_follow_up_date"] or ""),
    ("Next Follow-up Owner", lambda c: c["next_follow_up_owner"] or ""),
    ("Next Follow-up Note", lambda c: c["next_follow_up_note"] or ""),
    ("Created At", lambda c: _fmt_dt(c["created_at"])),
    ("Updated At", lambda c: _fmt_dt(c["updated_at"]))
]

ASSET_BASE_FIELDS = [
    ("ID", lambda a: a["id"]),
    ("Customer ID", lambda a: a["customer_id"]),
    ("Division", lambda a: a["division"] or ""),
    ("Plant", lambda a: a["plant"] or ""),
    ("Department", lambda a: a["department"] or ""),
    ("Asset", lambda a: a["name"] or ""),
    ("Asset Type", lambda a: a["asset_type"] or ""),
    ("Cleaning Frequency", lambda a: a["cleaning_frequency"] or ""),
    ("Observed Material", lambda a: a["observed_material"] or ""),
    ("Access Opening Type", lambda a: a["access_opening_type"] or ""),
    ("Equipment Nearby Possible", lambda a: _fmt_bool(a["can_place_equipment_nearby"])),
    ("Pain Point", lambda a: a["pain_point"] or ""),
    ("Created At", lambda a: _fmt_dt(a["created_at"]))
]

# (key, label, is_bool) - values come from asset.profile[key]. Mirrors
# frontend/src/data/surveyProfileFields.js's SURVEY_PROFILE_FIELDS/
# SURVEY_PROFILE_GROUPS exactly - see that file's header comment for
# the conditional-dewatering-group note (backend only ever writes
# those 15 keys when the source survey's dewatering_required was
# "Yes", so this group reads as "-" for every row otherwise).
SURVEY_PROFILE_GROUPS = [
    ("Site Profile — Customer Details", [
        ("nearest_hub", "Nearest Hub", False),
        ("urgency", "Urgency", False),
        ("survey_date", "Survey Date", False),
        ("surveyed_by", "Surveyed By", False),
        ("repeat_potential", "Repeat Potential", False)
    ]),
    ("Site Profile — Job Details", [
        ("cleaning_date", "Cleaning Date", False),
        ("material_ph_condition", "pH / Corrosiveness (Material)", False),
        ("sample_available", "Sample Available", False),
        ("temperature_range", "Temperature", False)
    ]),
    ("Site Profile — Sludge Details", [
        ("material_category", "Material Category", False),
        ("tank_type", "Tank Type", False),
        ("sludge_hardness", "Sludge Hardness", False),
        ("debris_level", "Debris Level", False),
        ("water_visibility", "Water Visibility", False),
        ("hazard_level", "Hazard Level", False),
        ("last_survey_id", "Last Survey ID", False),
        ("last_survey_date", "Last Survey Date", False)
    ]),
    ("Site Profile — Geometry", [
        ("tank_length", "Tank Length/Dia (m)", False),
        ("tank_width", "Tank Width (m)", False),
        ("tank_depth", "Sludge Depth (m)", False),
        ("opening_length", "Opening Length (mm)", False),
        ("opening_width", "Opening Width (mm)", False),
        ("opening_height", "Opening Height (mm)", False),
        ("height_from_ground", "Height From Ground (m)", False),
        ("drop_to_floor", "Drop To Floor (m)", False),
        ("vertical_lift", "Vertical Lift (m)", False)
    ]),
    ("Site Profile — Access & Setup", [
        ("hose_distance", "Hose Distance (m)", False),
        ("access_path_width", "Access Path Width (m)", False),
        ("access_support", "Access Support", False),
        ("customer_support", "Customer Support", False),
        ("access_type", "Access Type", False),
        ("equipment_nearby", "Equipment Nearby", False),
        ("scaffolding_needed", "Scaffolding Needed", True),
        ("crane_available", "Crane Available", True),
        ("tank_location", "Tank Location", False),
        ("setup_complexity", "Setup Complexity", False)
    ]),
    ("Site Profile — Safety", [
        ("power_available", "Power Available", False),
        ("water_available", "Water Available", True),
        ("air_supply_available", "Air Supply Available", False),
        ("confined_space", "Confined Space", True),
        ("ventilation_required", "Ventilation Required", True),
        ("gas_testing_required", "Gas Testing Required", True),
        ("ehs_restriction", "EHS Restriction", False),
        ("power_distance", "Power Distance (m)", False)
    ]),
    ("Site Profile — Pump Details", [
        ("abrasiveness", "Abrasiveness", False),
        ("pump_power_source", "Pump Power Source", False),
        ("discharge_medium", "Discharge Medium", False),
        ("disposal_responsibility", "Disposal Responsibility", False),
        ("discharge_point_distance", "Discharge Point Distance (m)", False),
        ("suction_depth", "Suction Depth (m)", False),
        ("discharge_distance", "Discharge Distance (m)", False),
        ("discharge_pit_dimension", "Discharge Pit Dimension", False)
    ]),
    ("Site Profile — Dewatering", [
        ("dewatering_required", "Dewatering Required", False),
        ("dewatering_volume", "Dewatering Volume (m³)", False),
        ("inlet_moisture", "Inlet Moisture %", False),
        ("target_final_moisture", "Target Final Moisture %", False),
        ("expected_final_form", "Expected Final Form", False),
        ("visible_free_water", "Visible Free Water", False),
        ("natural_settling", "Natural Settling Ability", False),
        ("oily_emulsified", "Oily / Emulsified", True),
        ("space_available", "Space for Bags / Holding", False),
        ("filtrate_route", "Filtrate Route Available", True),
        ("moisture_guarantee", "Final Moisture Guarantee", True),
        ("cake_handling_scope", "Cake Handling Scope", False),
        ("filtrate_route_detail", "Filtrate Route Detail", False),
        ("polymer_allowed", "Polymer Allowed", False),
        ("commitment", "Commitment", False)
    ]),
    ("Site Profile — Customer Insights", [
        ("customer_pain_point", "Customer Pain", False),
        ("shutdown_window", "Shutdown Window", False),
        ("current_method", "Current Method", False),
        ("budget_estimate", "Budget Estimate (INR)", False),
        ("decision_maker", "Decision Maker", False)
    ])
]


# Fields whose values tend to be long free text, read better wrapped
# across a taller row rather than one truncated line.
_WRAP_LABELS = {"Pain Point", "Next Follow-up Note"}


# ====================================
# DETAILS SHEET (single instance - Field | Value, matches the
# reference report's own two-column form exactly)
# ====================================

def _build_details_sheet(wb, customer):

    ws = wb.create_sheet(_sheet_name("Details", customer["company_name"]))

    total_cols = 2
    row = 1

    _title_row(ws, row, total_cols, "JANYU TECHNOLOGIES")
    row += 1

    _subtitle_row(ws, row, total_cols, f"Customer Profile — {customer['company_name']}")
    row += 1

    _section_band(ws, row, total_cols, "Customer Details")
    row += 1

    for label, accessor in CUSTOMER_FIELDS:
        _field_row(ws, row, label, [accessor(customer)], wrap=label in _WRAP_LABELS)
        row += 1

    ws.column_dimensions["A"].width = LABEL_COL_WIDTH
    ws.column_dimensions["B"].width = 46

    ws.freeze_panes = "A5"

    return ws


# ====================================
# ASSET SHEETS - one sheet per asset, not one shared sheet with an
# asset per column. Same Field | Value two-column form as the Details
# sheet, just repeated once per asset so each asset's own data (and
# its own conditional groups - see the Dewatering skip below) stands
# on its own rather than being crammed sideways alongside every other
# asset on file.
# ====================================

def _build_asset_sheet(wb, customer, asset):

    label = f"{asset['name'] or 'Asset'} #{asset['id']}"

    ws = wb.create_sheet(_sheet_name(label, customer["company_name"]))

    total_cols = 2
    row = 1

    _title_row(ws, row, total_cols, "JANYU TECHNOLOGIES")
    row += 1

    _subtitle_row(ws, row, total_cols, f"Asset & Site Profile — {label}")
    row += 1

    _section_band(ws, row, total_cols, "Asset Details")
    row += 1

    for field_label, accessor in ASSET_BASE_FIELDS:
        _field_row(ws, row, field_label, [accessor(asset)], wrap=field_label in _WRAP_LABELS)
        row += 1

    profile = asset.get("profile") or {}

    for group_title, fields in SURVEY_PROFILE_GROUPS:

        # This asset's last survey never wrote any of the 15 Dewatering
        # keys (sync_asset_profile_from_survey only does so when that
        # survey's own dewatering_required == "Yes") - skip the whole
        # group rather than showing it as 15 blank rows.
        if group_title == "Site Profile — Dewatering" and not profile.get("dewatering_required"):
            continue

        _section_band(ws, row, total_cols, group_title)
        row += 1

        for key, field_label, is_bool in fields:

            raw = profile.get(key)
            value = "" if raw is None else (_fmt_bool(raw) if is_bool else raw)

            _field_row(ws, row, field_label, [value])
            row += 1

    ws.column_dimensions["A"].width = LABEL_COL_WIDTH
    ws.column_dimensions["B"].width = VALUE_COL_WIDTH

    ws.freeze_panes = "A4"

    return ws


def _build_no_assets_sheet(wb, customer):

    ws = wb.create_sheet(_sheet_name("Assets", customer["company_name"]))

    total_cols = 2
    row = 1

    _title_row(ws, row, total_cols, "JANYU TECHNOLOGIES")
    row += 1

    _subtitle_row(ws, row, total_cols, f"Asset & Site Profile Register — {customer['company_name']}")
    row += 1

    _field_row(ws, row, "Assets on file", ["None yet."])

    ws.column_dimensions["A"].width = LABEL_COL_WIDTH
    ws.column_dimensions["B"].width = VALUE_COL_WIDTH

    return ws


# ====================================
# BUILD WORKBOOK
# ====================================

def build_customer_360_workbook(db, customer_id):

    customer = get_customer_detail_request(db, customer_id)

    if customer is None:
        return None

    wb = Workbook()
    wb.remove(wb.active)

    _build_details_sheet(wb, customer)

    assets = customer["assets"] or []

    if not assets:
        _build_no_assets_sheet(wb, customer)
    else:
        for asset in assets:
            _build_asset_sheet(wb, customer, asset)

    return wb, customer["company_name"]


def build_customer_360_workbook_bytes(db, customer_id):

    result = build_customer_360_workbook(db, customer_id)

    if result is None:
        return None, None

    wb, company_name = result

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    return buffer, company_name
