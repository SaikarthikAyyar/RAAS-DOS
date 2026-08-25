# ====================================
# IMPORTS
# ====================================

from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from backend.services.fleet_availability_service import _forecast_data


# ====================================
# STYLE CONSTANTS
# Matches the reference spreadsheet's own real look - blue header
# band, light-blue aggregate summary row, full-row green category
# bands, light-green equipment rows.
# ====================================

HEADER_FILL = PatternFill("solid", fgColor="4472C4")
HEADER_FONT = Font(color="FFFFFF", bold=True, size=10)

SUMMARY_FILL = PatternFill("solid", fgColor="DCE6F1")
SUMMARY_FONT = Font(bold=True, size=10)

CATEGORY_FILL = PatternFill("solid", fgColor="6AA84F")
CATEGORY_FONT = Font(color="FFFFFF", bold=True, size=10.5)

ROW_FILL = PatternFill("solid", fgColor="E2EFDA")

THIN_BORDER = Border(*(Side(style="thin", color="D9D9D9"),) * 4)

CURRENCY_FORMAT = '"₹"#,##0.00;("₹"#,##0.00)'

FIXED_HEADERS = ["Sr No", "Fleet Unit", "Machine", "Current Location", "Status"]


# ====================================
# BUILD WORKBOOK
# ====================================

def build_forecast_workbook(db, weeks=13):

    data = _forecast_data(db, weeks)

    wb = Workbook()
    ws = wb.active
    ws.title = "3-Month Forecast"

    total_cols = len(FIXED_HEADERS) + len(data["months"]) + len(data["weeks"])

    # ---- header row ----
    col = 1

    for label in FIXED_HEADERS:
        cell = ws.cell(row=1, column=col, value=label)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        col += 1

    for month in data["months"]:
        cell = ws.cell(row=1, column=col, value=month["label"])
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")
        col += 1

    for week in data["weeks"]:
        cell = ws.cell(row=1, column=col, value=f"{week['label']}\n{week['range_label']}")
        cell.fill = HEADER_FILL
        cell.font = Font(color="FFFFFF", bold=True, size=8.5)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        col += 1

    ws.row_dimensions[1].height = 32

    # ---- aggregate "Billed Value" summary row ----
    row = 2

    ws.cell(row=row, column=2, value="Billed Value")

    for c in range(1, total_cols + 1):
        ws.cell(row=row, column=c).fill = SUMMARY_FILL
        ws.cell(row=row, column=c).font = SUMMARY_FONT

    month_start_col = len(FIXED_HEADERS) + 1

    for i, month in enumerate(data["monthly_billed_value"]):
        cell = ws.cell(row=row, column=month_start_col + i, value=month["total"] or None)
        cell.number_format = CURRENCY_FORMAT
        cell.alignment = Alignment(horizontal="center")

    row += 1

    # blank spacer row
    row += 1

    # ---- category-grouped equipment rows ----
    sr_no = 1
    current_category = None

    week_start_col = len(FIXED_HEADERS) + len(data["months"]) + 1

    for unit_row in data["rows"]:

        if unit_row["category"] != current_category:

            current_category = unit_row["category"]

            band_cell = ws.cell(row=row, column=1, value=current_category)
            band_cell.font = CATEGORY_FONT

            for c in range(1, total_cols + 1):
                ws.cell(row=row, column=c).fill = CATEGORY_FILL

            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=len(FIXED_HEADERS))

            row += 1

        values = [
            sr_no,
            f"{unit_row['fleet_code']} - {unit_row['fleet_name']}",
            unit_row["machine_code"] or "-",
            unit_row["current_location"] or "-",
            unit_row["utilization"]
        ]

        for i, value in enumerate(values, start=1):
            cell = ws.cell(row=row, column=i, value=value)
            cell.fill = ROW_FILL
            cell.border = THIN_BORDER

        for i, monthly_value in enumerate(unit_row["monthly_values"]):
            cell = ws.cell(row=row, column=month_start_col + i, value=monthly_value or None)
            cell.number_format = CURRENCY_FORMAT
            cell.fill = ROW_FILL
            cell.border = THIN_BORDER
            cell.alignment = Alignment(horizontal="center")

        for i, occupant in enumerate(unit_row["cells"]):
            cell = ws.cell(row=row, column=week_start_col + i, value=occupant or None)
            cell.fill = ROW_FILL
            cell.border = THIN_BORDER
            cell.alignment = Alignment(horizontal="center")
            cell.font = Font(size=9)

        sr_no += 1
        row += 1

    # ---- column widths ----
    ws.column_dimensions["A"].width = 6
    ws.column_dimensions["B"].width = 32
    ws.column_dimensions["C"].width = 18
    ws.column_dimensions["D"].width = 18
    ws.column_dimensions["E"].width = 9

    for i in range(len(data["months"])):
        ws.column_dimensions[get_column_letter(month_start_col + i)].width = 14

    for i in range(len(data["weeks"])):
        ws.column_dimensions[get_column_letter(week_start_col + i)].width = 16

    ws.freeze_panes = "F2"

    return wb


def build_forecast_workbook_bytes(db, weeks=13):

    wb = build_forecast_workbook(db, weeks)

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    return buffer
