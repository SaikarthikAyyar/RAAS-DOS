# ====================================
# SHARED XLSX STYLE
# Common look-and-feel for every server-side-generated .xlsx export in
# this app (title banner, blue subtitle band, section bands, bordered
# grid) - matches the field-visit-report look this business already
# uses on paper. First built for the Customer 360 export
# (backend/reporting/customer_360_xlsx.py); extracted here so every
# other export (Business Masters' "Export current tab") renders with
# the exact same design instead of a second, hand-copied version of
# it silently drifting out of sync.
#
# The client-side SheetJS export every one of these used to be built
# from (the free/community `xlsx` npm package) can't apply cell
# fills/fonts/merges at all - that's the reason this whole family of
# exports moved server-side, generated with openpyxl instead.
# ====================================

import re

from openpyxl.styles import Font, PatternFill, Alignment, Border, Side


# ====================================
# STYLE CONSTANTS
# ====================================

TITLE_FONT = Font(bold=True, size=14, color="1F3864")
TITLE_BORDER = Border(bottom=Side(style="medium", color="1F3864"))

SUBTITLE_FILL = PatternFill("solid", fgColor="1F3864")
SUBTITLE_FONT = Font(bold=True, size=11, color="FFFFFF")

SECTION_FILL = PatternFill("solid", fgColor="8EA9DB")
SECTION_FONT = Font(bold=True, size=10.5, color="1F3864")

COLUMN_HEADER_FILL = PatternFill("solid", fgColor="D9E2F3")
COLUMN_HEADER_FONT = Font(bold=True, size=10)

LABEL_FILL = PatternFill("solid", fgColor="F2F2F2")
LABEL_FONT = Font(bold=True, size=10)

VALUE_FONT = Font(size=10)

THIN_BORDER = Border(*(Side(style="thin", color="BFBFBF"),) * 4)

LABEL_COL_WIDTH = 32
VALUE_COL_WIDTH = 30

# Longer free-text fields read better with a taller wrapped row instead
# of one long unreadable line.
WRAP_ROW_HEIGHT = 42


# ====================================
# SHEET-NAME SAFETY (Excel's 31-char cap, and a handful of characters
# Excel rejects outright in a sheet name)
# ====================================

def sheet_name(suffix, prefix=""):

    clean_prefix = re.sub(r'[\[\]:*?/\\]', "", prefix or "")
    clean_suffix = re.sub(r'[\[\]:*?/\\]', "", suffix or "")

    if not clean_prefix:
        return clean_suffix[:31]

    tail = f" - {clean_suffix}"
    truncated_prefix = clean_prefix[:max(31 - len(tail), 0)]

    return f"{truncated_prefix}{tail}"


# ====================================
# SHARED ROW HELPERS
# ====================================

def title_row(ws, row, total_cols, text):

    cell = ws.cell(row=row, column=1, value=text)
    cell.font = TITLE_FONT
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border = TITLE_BORDER

    for c in range(2, total_cols + 1):
        ws.cell(row=row, column=c).border = TITLE_BORDER

    if total_cols > 1:
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=total_cols)

    ws.row_dimensions[row].height = 24


def subtitle_row(ws, row, total_cols, text):

    cell = ws.cell(row=row, column=1, value=text)
    cell.fill = SUBTITLE_FILL
    cell.font = SUBTITLE_FONT
    cell.alignment = Alignment(horizontal="center", vertical="center")

    for c in range(2, total_cols + 1):
        ws.cell(row=row, column=c).fill = SUBTITLE_FILL

    if total_cols > 1:
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=total_cols)

    ws.row_dimensions[row].height = 20


def section_band(ws, row, total_cols, text):

    cell = ws.cell(row=row, column=1, value=text)
    cell.fill = SECTION_FILL
    cell.font = SECTION_FONT
    cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)

    for c in range(2, total_cols + 1):
        ws.cell(row=row, column=c).fill = SECTION_FILL

    if total_cols > 1:
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=total_cols)

    ws.row_dimensions[row].height = 20


def field_row(ws, row, label, values, wrap=False):

    label_cell = ws.cell(row=row, column=1, value=label)
    label_cell.fill = LABEL_FILL
    label_cell.font = LABEL_FONT
    label_cell.border = THIN_BORDER
    label_cell.alignment = Alignment(vertical="center", wrap_text=True)

    for i, value in enumerate(values, start=2):
        cell = ws.cell(row=row, column=i, value=value if value not in (None, "") else None)
        cell.font = VALUE_FONT
        cell.border = THIN_BORDER
        cell.alignment = Alignment(vertical="center", wrap_text=wrap)

    if wrap:
        ws.row_dimensions[row].height = WRAP_ROW_HEIGHT


# ====================================
# TABULAR ROW HELPERS (real column-headers-then-records tables, as
# opposed to the vertical Field | Value form the helpers above build -
# what "Export current tab" needs, since it's a raw table dump rather
# than a profile-style form)
# ====================================

def table_header_row(ws, row, headers):

    for i, header in enumerate(headers, start=1):

        cell = ws.cell(row=row, column=i, value=header)
        cell.fill = COLUMN_HEADER_FILL
        cell.font = COLUMN_HEADER_FONT
        cell.border = THIN_BORDER
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    ws.row_dimensions[row].height = 26


def table_data_row(ws, row, values):

    for i, value in enumerate(values, start=1):

        cell = ws.cell(row=row, column=i, value=value if value not in (None, "") else None)
        cell.font = VALUE_FONT
        cell.border = THIN_BORDER
        cell.alignment = Alignment(vertical="center")
